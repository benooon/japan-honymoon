#!/usr/bin/env python3
"""
Rebuilds the DATA.expenses / DATA.expenses_meta block inside index.html from
the raw credit-card statement exports in data/statements/*.xlsx.

Usage:
    1. Drop new statement exports (.xlsx) from any card into data/statements/.
       Supported formats: Leumi Card "transactiondetails_export" and
       Isracard/CAL "פירוט עסקאות" personal-area exports.
    2. Run:  python3 scripts/build_expenses.py
    3. Commit the updated index.html (and the new file(s) under data/statements/).

Re-running is idempotent and safe to do after adding more files — every run
re-parses everything in data/statements/ from scratch and dedupes exact
repeats (same date + merchant + amount) that show up in overlapping exports.
"""
import json
import re
import sys
from pathlib import Path
from datetime import datetime

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
STATEMENTS_DIR = ROOT / "data" / "statements"
INDEX_HTML = ROOT / "index.html"

JAPAN_CITY_HINTS = [
    "JPN", "TOKYO", "TOUKIYOUTO", "KYOTO", "OSAKA", "GINZA", "SHIBUYA",
    "HARAJUKU", "AKIHABARA", "ROPPONGI", "SAWANDO", "HAKONE",
]
ALWAYS_TRIP_HINTS = ["BOOKING.COM HOTEL", "BKG*BOOKING.COM", "MOBIMATTER"]

CATEGORY_RULES = [
    ("\U0001F3E8 מלונות", [  # 🏨 מלונות
        "BOOKING.COM", "BKG*BOOKING", "GINZA HOTEL", "GORA KANSUIRO",
        "GOURAKANSUIROU", "RYOKAN",
    ]),
    ("\U0001F35C אוכל ומסעדות", [  # 🍜 אוכל ומסעדות
        "STARBUCKS", "MATCHA", "BILLS ", "COFFEE", "KOFFEE", "CAFE",
        "HAND ROLL", "YAKINIKU", "NIKUNOHANAMASA", "KYUYAMUTEI", "JINZU",
        "SIDEWALK", "ONEBYONE", "BONGEN", "SEVEN-ELEVEN", "LAWSON", "DYDO",
        "APFR", "RIOSHIMOKITAZAWA", "REATENDON", "GOSENJAKU", "KOSHUKYOUDO",
        "GINZAKAGARI", "TABLECHECK", "QWANG", "PST ROPPONGI", "BAITEN",
        "KAJIWARA KITCHEN", "GRILL", "BREERY", "BREWERY",
    ]),
    ("\U0001F6CD️ קניות", [  # 🛍️ קניות
        "UNIQLO", "PUMA", "NIKE", "ABC-MART", "EDWIN", "LOFT",
        "MATSUMOTOKIYOSHI", "OBEY", "DONQUIJOTE", "MUJIRUSHIRYOHIN",
        "OMOTESANDO HILLS", "TRADING CARD SHOP", "SOUVENIR", "SHOTEN",
        "KOSAI", "AKIHABARA",
    ]),
    ("\U0001F3A8 פעילויות", [  # 🎨 פעילויות
        "MUSEUM", "CHOKOKUNOMORI", "OPEN-AIR",
    ]),
    ("\U0001F686 תחבורה", [  # 🚆 תחבורה
        "SUICA", "JR EAST", "ALPICO", "UBER", "TYUUSYARYOUKIN",
        "EMOT(RAILWAY)",
    ]),
    ("\U0001F4F1 תקשורת", [  # 📱 תקשורת
        "MOBIMATTER",
    ]),
    ("\U0001F4B5 מזומן שנמשך", [  # 💵 מזומן שנמשך
        "SEVEN BANK",
    ]),
]
LODGING_CATEGORY = CATEGORY_RULES[0][0]
CASH_CATEGORY = CATEGORY_RULES[-1][0]
MISC_CATEGORY = "\U0001F4CB שונות"  # 📋 שונות


def categorize(merchant: str) -> str:
    m = merchant.upper()
    for cat, keywords in CATEGORY_RULES:
        if any(k.upper() in m for k in keywords):
            return cat
    return MISC_CATEGORY


def is_japan_trip(currency, merchant: str) -> bool:
    m = merchant.upper()
    if currency == "¥":  # ¥
        return True
    if any(h in m for h in JAPAN_CITY_HINTS):
        return True
    if any(h in m for h in ALWAYS_TRIP_HINTS):
        return True
    return False


def parse_date(s):
    s = str(s).strip()
    for fmt in ("%d-%m-%Y", "%d.%m.%y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def clean_merchant(name: str) -> str:
    return re.sub(r"\s{2,}", " ", str(name)).strip()


def parse_leumi(ws_by_title):
    rows = []
    for title in ("עסקאות במועד החיוב",
                  "עסקאות חו\"ל ומט\"ח"):
        ws = ws_by_title.get(title)
        if not ws:
            continue
        header_seen = False
        for row in ws.iter_rows(values_only=True):
            if not header_seen:
                if row and row[0] == "תאריך עסקה":
                    header_seen = True
                continue
            if not row or row[0] is None:
                continue
            date_raw, merchant, category, *_rest = row
            amount, currency = row[5], row[6]
            if not isinstance(amount, (int, float)):
                continue
            d = parse_date(date_raw)
            if not d:
                continue
            merchant = clean_merchant(merchant)
            if not is_japan_trip(currency, merchant):
                continue
            rows.append({
                "date": d.isoformat(),
                "merchant": merchant,
                "nis": round(float(amount), 2),
                "card": "Leumi 9420",
                "cat": categorize(merchant),
            })
    return rows


def parse_isracard(ws, card_label):
    rows = []
    section = None
    header = None
    pending = []  # buffered rows for sections without a per-row ILS amount

    def flush_pending(ils_total):
        """The 'not yet posted' section only states a per-line JPY/orig amount
        plus one section-wide ILS total, so distribute it pro-rata by amount."""
        if not pending:
            return
        ils_rows = [p for p in pending if p["orig_currency"] == "₪"]
        fx_rows = [p for p in pending if p["orig_currency"] != "₪"]
        ils_sum = sum(p["orig_amount"] for p in ils_rows)
        fx_sum = sum(p["orig_amount"] for p in fx_rows)
        remaining = (ils_total or 0) - ils_sum
        rate = (remaining / fx_sum) if fx_sum else 0
        for p in pending:
            amount = p["orig_amount"] if p["orig_currency"] == "₪" else p["orig_amount"] * rate
            add_row(p["date"], p["merchant"], amount, p["orig_currency"], None)
        pending.clear()

    def add_row(d, merchant, amount, orig_currency, discount):
        if not isinstance(amount, (int, float)):
            return
        if not is_japan_trip(orig_currency, merchant):
            return
        rows.append({
            "date": d.isoformat(),
            "merchant": merchant,
            "nis": round(float(amount), 2),
            "card": card_label,
            "cat": categorize(merchant),
            "fee_discount": discount,
        })

    for row in ws.iter_rows(values_only=True):
        first = row[0]
        if first in ("עסקאות שטרם נקלטו",
                     "עסקאות למועד חיוב",
                     "עסקאות בחיוב מחוץ למועד"):
            flush_pending(None)
            section = first
            header = None
            continue
        if first == "תאריך רכישה":
            header = row
            continue
        if section is None or header is None:
            continue
        if first is None and row[1] and str(row[1]).startswith('סה"כ'):
            flush_pending(row[2] if isinstance(row[2], (int, float)) else None)
            continue
        if first is None or str(first).startswith("תנאים") or str(first).startswith("עסקאות שבוצעו"):
            flush_pending(None)
            section = None
            continue
        d = parse_date(first)
        if not d:
            continue
        merchant = clean_merchant(row[1])
        orig_amount, orig_currency = row[2], row[3]
        if not isinstance(orig_amount, (int, float)):
            continue
        # "billed" sections carry the ILS-converted amount in cols 4/5 already;
        # the "not yet posted" section only has the original-currency amount
        # and needs to be prorated once its section total row is seen.
        if len(row) > 5 and isinstance(row[4], (int, float)) and row[5]:
            discount = None
            if len(row) > 7 and row[7]:
                m = re.search(r"הנחה ₪([\d.]+)", str(row[7]))
                if m:
                    discount = float(m.group(1))
            add_row(d, merchant, row[4], orig_currency, discount)
        else:
            pending.append({
                "date": d, "merchant": merchant,
                "orig_amount": orig_amount, "orig_currency": orig_currency,
            })
    flush_pending(None)
    return rows


def load_workbook_rows(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    titles = {ws.title: ws for ws in wb.worksheets}
    if "עסקאות במועד החיוב" in titles:
        return parse_leumi(titles)
    if "פירוט עסקאות" in titles:
        ws = titles["פירוט עסקאות"]
        # card label + last 4 digits sit a few rows down, e.g. "MC דירקט - 2275"
        card_label = path.stem
        for row in ws.iter_rows(values_only=True, max_row=8):
            if row[0] and " - " in str(row[0]):
                card_label = str(row[0]).replace("‫", "").replace("‬", "")
                break
        return parse_isracard(ws, card_label)
    print(f"WARNING: unrecognised workbook format, skipping: {path.name}", file=sys.stderr)
    return []


def main():
    if not STATEMENTS_DIR.exists():
        print(f"No statements dir at {STATEMENTS_DIR}", file=sys.stderr)
        sys.exit(1)

    all_rows = []
    for path in sorted(STATEMENTS_DIR.glob("*.xlsx")):
        all_rows.extend(load_workbook_rows(path))

    # dedupe exact repeats across overlapping statement exports
    seen = set()
    rows = []
    for r in all_rows:
        key = (r["date"], r["merchant"], r["nis"], r["card"])
        if key in seen:
            continue
        seen.add(key)
        rows.append(r)

    rows.sort(key=lambda r: r["date"])

    spend_rows = [r for r in rows if r["cat"] != CASH_CATEGORY]
    cash_rows = [r for r in rows if r["cat"] == CASH_CATEGORY]
    # hotels are booked/paid in lump sums (and tracked separately in the
    # planned-budget tab) so they'd swamp a "daily spend" average — keep
    # them in the category totals but out of the daily-average calculation.
    daily_rows = [r for r in spend_rows if r["cat"] != LODGING_CATEGORY]

    total_spend = round(sum(r["nis"] for r in spend_rows), 2)
    lodging_total = round(sum(r["nis"] for r in spend_rows if r["cat"] == LODGING_CATEGORY), 2)
    daily_spend_total = round(sum(r["nis"] for r in daily_rows), 2)
    total_cash = round(sum(r["nis"] for r in cash_rows), 2)
    total_fee_savings = round(sum(r.get("fee_discount") or 0 for r in rows), 2)

    dates = sorted({r["date"] for r in spend_rows})
    if dates:
        span_days = (datetime.fromisoformat(dates[-1]).date() - datetime.fromisoformat(dates[0]).date()).days + 1
    else:
        span_days = 0
    daily_avg = round(daily_spend_total / span_days, 2) if span_days else 0

    by_cat = {}
    for r in spend_rows:
        by_cat[r["cat"]] = round(by_cat.get(r["cat"], 0) + r["nis"], 2)

    expenses_meta = {
        "total_nis": total_spend,
        "lodging_nis": lodging_total,
        "daily_spend_nis": daily_spend_total,
        "cash_withdrawn_nis": total_cash,
        "fee_savings_nis": total_fee_savings,
        "first_date": dates[0] if dates else None,
        "last_date": dates[-1] if dates else None,
        "days_span": span_days,
        "daily_avg_nis": daily_avg,
        "daily_avg_per_person_nis": round(daily_avg / 2, 2) if daily_avg else 0,
        "by_category": [{"cat": k, "nis": v} for k, v in sorted(by_cat.items(), key=lambda kv: -kv[1])],
        "generated_from": [p.name for p in sorted(STATEMENTS_DIR.glob("*.xlsx"))],
    }

    expenses = [
        {k: v for k, v in r.items() if k != "fee_discount"}
        for r in rows
    ]

    html = INDEX_HTML.read_text(encoding="utf-8")
    m = re.search(r"const DATA = (\{[\s\S]*?\});\n", html)
    if not m:
        print("FAIL: could not find DATA object in index.html", file=sys.stderr)
        sys.exit(1)
    data = json.loads(m.group(1))
    data["expenses"] = expenses
    data["expenses_meta"] = expenses_meta

    new_block = "const DATA = " + json.dumps(data, ensure_ascii=False, indent=1) + ";\n"
    html = html[:m.start()] + new_block + html[m.end():]
    INDEX_HTML.write_text(html, encoding="utf-8")

    print(f"OK — {len(expenses)} Japan transactions "
          f"({len(spend_rows)} spend + {len(cash_rows)} cash withdrawals) "
          f"from {len(list(STATEMENTS_DIR.glob('*.xlsx')))} statement file(s)")
    print(f"Total spent in Japan so far: ₪{total_spend:,.2f} "
          f"over {span_days} days ({dates[0] if dates else '-'} .. {dates[-1] if dates else '-'})")
    print(f"  of which lodging (excluded from daily average): ₪{lodging_total:,.2f}")
    print(f"Daily average (excl. lodging): ₪{daily_avg:,.2f}/day (₪{expenses_meta['daily_avg_per_person_nis']:,.2f}/person/day)")
    print(f"Cash withdrawn (not itemised): ₪{total_cash:,.2f}")
    print(f"Foreign-currency fee discounts (מועדון הנחות) saved: ₪{total_fee_savings:,.2f}")


if __name__ == "__main__":
    main()
